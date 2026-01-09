"""
核心比较器：管理工作簿、选择单元格、直接比较、公式验证、导出结果（基本）
"""
from core.excel_reader import load_workbook_all_sheets
from core.string_comparator import StringComparator
from core.validator import validate_formula
from core.rule_engine import RuleEngine
import pandas as pd
import re
import logging
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# 配置日志记录
logging.basicConfig(level=logging.DEBUG, 
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler("app.log"),
                        logging.StreamHandler()
                    ])
logger = logging.getLogger(__name__)

class ExcelComparator:
    def __init__(self):
        """
        初始化Excel比较器
        
        创建工作簿字典、字符串比较器和规则引擎实例
        工作簿字典结构：{alias: {'path': 文件路径, 'sheets': {工作表名: DataFrame}}}
        """
        self.workbooks = {}  # alias -> { 'path':..., 'sheets': {name:DataFrame} }
        self.string_comparator = StringComparator()
        self.rule_engine = RuleEngine()

    def load_workbook(self, filepath, alias=None):
        """
        加载Excel工作簿并存储
        
        参数:
            filepath: Excel文件路径
            alias: 工作簿别名，默认为文件路径
        
        将工作簿的所有工作表加载为DataFrame，并存储在工作簿字典中
        
        返回:
            dict: 工作表名称到DataFrame的映射
            
        异常:
            Exception: 如果文件加载失败，会捕获并重新抛出异常
        """
        alias = alias or filepath
        logger.info(f"加载工作簿: {filepath}，别名为: {alias}")
        try:
            sheets = load_workbook_all_sheets(filepath)
            self.workbooks[alias] = {
                'path': filepath,
                'sheets': sheets
            }
            logger.info(f"工作簿加载成功，包含 {len(sheets)} 个工作表: {list(sheets.keys())}")
            return sheets
        except Exception as e:
            logger.error(f"加载工作簿失败: {filepath}，错误: {str(e)}")
            raise Exception(f"无法加载工作簿 {filepath}: {str(e)}") from e
    def list_sheets(self, alias):
        """
        获取指定工作簿的所有工作表名称列表
        
        参数:
            alias: 工作簿别名
            
        返回:
            list: 工作表名称列表，如果工作簿不存在则返回空列表
        """
        if alias not in self.workbooks:
            return []
        return list(self.workbooks[alias]['sheets'].keys())

    def get_sheet_dataframe(self, alias, sheet_name):
        """
        获取指定工作表的数据框
        
        参数:
            alias: 工作簿别名
            sheet_name: 工作表名称
            
        返回:
            DataFrame: 工作表的数据框，已重置索引为0-based
            
        异常:
            ValueError: 当工作簿或工作表不存在时抛出
        """
        logger.info(f"获取工作表数据框: 工作簿={alias}，工作表={sheet_name}")
        if alias not in self.workbooks:
            logger.error(f"未加载工作簿: {alias}")
            raise ValueError(f"未加载工作簿: {alias}")
        sheets = self.workbooks[alias]['sheets']
        if sheet_name not in sheets:
            logger.error(f"工作表 {sheet_name} 不存在")
            raise ValueError(f"工作表 {sheet_name} 不存在")
        df = sheets[sheet_name].copy()
        # reset index to simple 0..n-1 to align with model
        df = df.reset_index(drop=True)
        logger.info(f"成功获取工作表数据框，形状: {df.shape}")
        return df

    cell_ref_re = re.compile(r'^([A-Za-z]+)(\d+)$')

    @staticmethod
    def col_letters_to_index(col_letters):
        """A -> 0, B -> 1, AA -> 26"""
        col_letters = col_letters.upper()
        idx = 0
        for ch in col_letters:
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx - 1

    def parse_range(self, rng):
        """
        支持单元格或矩形:
        - "A1" -> (col0,row0, col0,row0)
        - "A1:C10" -> (col0,row0, col2,row9)
        1-based rows in Excel -> we convert to 0-based
        """
        parts = rng.split(':')
        if len(parts) == 1:
            m = self.cell_ref_re.match(parts[0])
            if not m:
                raise ValueError("无效单元格地址")
            c = self.col_letters_to_index(m.group(1))
            r = int(m.group(2)) - 1
            return c, r, c, r
        elif len(parts) == 2:
            m1 = self.cell_ref_re.match(parts[0])
            m2 = self.cell_ref_re.match(parts[1])
            if not m1 or not m2:
                raise ValueError("无效单元格范围")
            c1 = self.col_letters_to_index(m1.group(1))
            r1 = int(m1.group(2)) - 1
            c2 = self.col_letters_to_index(m2.group(1))
            r2 = int(m2.group(2)) - 1
            # normalize
            return min(c1,c2), min(r1,r2), max(c1,c2), max(r1,r2)
        else:
            raise ValueError("不支持的范围格式")

    def select_cells(self, workbook_alias, sheet_name, rng):
        """
        返回 pandas.DataFrame 对应范围（如果超出 sheet 大小，返回可用交集）
        """
        df = self.get_sheet_dataframe(workbook_alias, sheet_name)
        c1, r1, c2, r2 = self.parse_range(rng)
        # pandas uses columns as names; we convert by position
        max_cols = df.shape[1]
        max_rows = df.shape[0]
        # if sheet has no columns, return empty df with 0 cols
        if max_cols == 0:
            return pd.DataFrame()
        c1 = max(0, c1); c2 = min(max_cols - 1, c2)
        r1 = max(0, r1); r2 = min(max_rows - 1, r2)
        cols = list(df.columns[c1:c2+1])
        sub = df.loc[r1:r2, cols].reset_index(drop=True)
        # ensure column names are present
        return sub

    def compare_direct(self, df1, df2, options=None):
        """
        逐单元格比较两个DataFrame
        
        参数:
            df1: 第一个DataFrame
            df2: 第二个DataFrame
            options: 比较选项字典，支持以下键:
                - tolerance: 数值比较容差，默认0
                - ignore_case: 字符串比较是否忽略大小写，默认False
        
        返回:
            tuple: (result_df, result_map)
                - result_df: 比较结果DataFrame，以两个DataFrame的最大行列数为基准
                - result_map: 字典，键为(r,c)坐标，值为比较状态('equal'/'diff'/'empty')
                
        比较规则:
            - 数值比较：计算差值，在容差范围内视为相等
            - 字符串比较：根据ignore_case选项进行比较
            - 空值比较：两个空值视为相等
            - 不同类型比较：转换为字符串后比较
        """
        logger.debug(f"compare_direct - 输入df1类型: {type(df1)}, df1形状: {df1.shape}")
        logger.debug(f"compare_direct - 输入df2类型: {type(df2)}, df2形状: {df2.shape}")
        logger.debug(f"compare_direct - df1内容:\n{df1}")
        logger.debug(f"compare_direct - df2内容:\n{df2}")
        
        options = options or {}
        tol = float(options.get('tolerance', 0))
        ignore_case = bool(options.get('ignore_case', False))
        logger.debug(f"compare_direct - 比较选项: tolerance={tol}, ignore_case={ignore_case}")

        # 确定结果表的形状
        rows = max(df1.shape[0], df2.shape[0])
        cols = max(df1.shape[1], df2.shape[1])
        logger.debug(f"compare_direct - 结果表形状: rows={rows}, cols={cols}")

        # 确保结果表有足够的列
        col_names = []
        for i in range(cols):
            # 优先使用df1的列名，其次是df2，最后是通用名
            if i < df1.shape[1]:
                name = str(df1.columns[i])
            elif i < df2.shape[1]:
                name = str(df2.columns[i])
            else:
                name = f"COL_{i}"
            col_names.append(name)
        logger.debug(f"compare_direct - 列名: {col_names}")

        # 构建结果数据和映射
        result_vals = []
        result_map = {}
        
        for r in range(rows):
            row_vals = []
            for c in range(cols):
                # 获取单元格值
                v1 = df1.iloc[r, c] if r < df1.shape[0] and c < df1.shape[1] else None
                v2 = df2.iloc[r, c] if r < df2.shape[0] and c < df2.shape[1] else None
                
                logger.debug(f"compare_direct - 单元格({r},{c}) - v1: {v1}, 类型: {type(v1)}, v2: {v2}, 类型: {type(v2)}")
                
                # 处理空值情况
                is_v1_na = pd.isna(v1)
                is_v2_na = pd.isna(v2)
                
                logger.debug(f"compare_direct - 单元格({r},{c}) - is_v1_na: {is_v1_na}, is_v2_na: {is_v2_na}")
                
                # 默认状态和显示值
                status = 'empty'
                display = v1 if not is_v1_na else v2
                
                if is_v1_na and is_v2_na:
                    status = 'equal'
                else:
                    try:
                        # 数值比较
                        if isinstance(v1, (int, float)) and isinstance(v2, (int, float)):
                            logger.debug(f"compare_direct - 单元格({r},{c}) - 数值比较")
                            diff = abs(v1 - v2)
                            status = 'equal' if diff <= tol else 'diff'
                            logger.debug(f"compare_direct - 单元格({r},{c}) - 差值: {diff}, 容差: {tol}, 结果: {status}")
                        else:
                            # 转换为字符串比较
                            logger.debug(f"compare_direct - 单元格({r},{c}) - 字符串比较")
                            s1 = str(v1) if not is_v1_na else ''
                            s2 = str(v2) if not is_v2_na else ''
                            logger.debug(f"compare_direct - 单元格({r},{c}) - s1: {s1}, s2: {s2}")
                            
                            if ignore_case:
                                s1 = s1.lower()
                                s2 = s2.lower()
                                logger.debug(f"compare_direct - 单元格({r},{c}) - 忽略大小写后: s1={s1}, s2={s2}")
                            
                            status = 'equal' if s1 == s2 else 'diff'
                            logger.debug(f"compare_direct - 单元格({r},{c}) - 比较结果: {status}")
                    except Exception as e:
                        # 如果比较失败，标记为差异
                        logger.error(f"compare_direct - 单元格({r},{c}) - 比较出错: {e}")
                        status = 'diff'
                
                row_vals.append(display)
                result_map[(r, c)] = status
                logger.debug(f"compare_direct - 单元格({r},{c}) - 最终状态: {status}")
            
            result_vals.append(row_vals)
            logger.debug(f"compare_direct - 行{r}处理完成")

        # 创建结果DataFrame
        result_df = pd.DataFrame(result_vals, columns=col_names)
        return result_df, result_map

    def validate_formula(self, cells_dict, formula, expected_value, options=None):
        options = options or {}
        tol = float(options.get('tolerance', 0.0))
        return validate_formula(cells_dict, formula, expected_value, tolerance=tol)

    def export_results(self, result_df, output_path, format='excel'):
        """
        导出比较结果到文件
        
        参数:
            result_df: 比较结果DataFrame
            output_path: 输出文件路径
            format: 导出格式，支持'excel'和'csv'，默认为'excel'
            
        返回:
            bool: 导出是否成功
        """
        logger.info(f"开始导出结果到: {output_path}，格式: {format}")
        try:
            if format == 'excel':
                result_df.to_excel(output_path, index=False)
                logger.info(f"Excel文件导出成功: {output_path}")
            elif format == 'csv':
                result_df.to_csv(output_path, index=False)
                logger.info(f"CSV文件导出成功: {output_path}")
            else:
                logger.error(f"不支持的导出格式: {format}")
                raise ValueError(f"不支持的导出格式: {format}")
            return True
        except Exception as e:
            logger.error(f"导出结果失败: {str(e)}")
            return False
    
    def export_with_highlights(self, df, output_path, failed_cells=None, passed_cells=None):
        """
        导出带有颜色标记的表格到Excel文件
        
        参数:
            df: 要导出的数据框
            output_path: 输出文件路径
            failed_cells: 失败的单元格列表，格式为[(row1, col1), (row2, col2)]
            passed_cells: 通过的单元格列表，格式为[(row1, col1), (row2, col2)]
            
        返回:
            bool: 导出是否成功
        """
        logger.info(f"开始导出带有颜色标记的表格到: {output_path}")
        try:
            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active
            
            # 创建填充样式
            passed_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # 蓝色
            failed_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')  # 红色
            
            # 将DataFrame数据写入工作表
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                ws.append(row)
            
            # 标记失败的单元格
            if failed_cells:
                for row_idx, col_idx in failed_cells:
                    # 在Excel中，行和列都是从1开始的，且header占用了第一行
                    excel_row = row_idx + 2  # +1是因为DataFrame索引从0开始，+1是因为有header
                    excel_col = col_idx + 1  # 列索引从0开始
                    cell = ws.cell(row=excel_row, column=excel_col)
                    cell.fill = failed_fill
            
            # 标记通过的单元格
            if passed_cells:
                for row_idx, col_idx in passed_cells:
                    # 在Excel中，行和列都是从1开始的，且header占用了第一行
                    excel_row = row_idx + 2  # +1是因为DataFrame索引从0开始，+1是因为有header
                    excel_col = col_idx + 1  # 列索引从0开始
                    cell = ws.cell(row=excel_row, column=excel_col)
                    cell.fill = passed_fill
            
            # 保存文件
            wb.save(output_path)
            logger.info(f"带有颜色标记的Excel文件导出成功: {output_path}")
            return True
        except Exception as e:
            logger.error(f"导出带有颜色标记的表格失败: {str(e)}")
            return False

    def add_rule(self, rule):
        """
        添加自定义规则到规则引擎
        
        参数:
            rule: 规则字符串，如 "A1 + B1 = C1" 或 "FILE1:A1 = FILE2:A1"
        """
        self.rule_engine.add_rule(rule)
    
    def clear_rules(self):
        """
        清除所有规则
        """
        self.rule_engine.clear_rules()
    
    def get_rules(self):
        """
        获取当前所有规则
        
        返回:
            list: 规则列表
        """
        return self.rule_engine.rules
    
    def validate_with_dataframes(self, df1, df2):
        """
        使用数据帧验证规则（支持单表或跨文件比较）
        
        参数:
            df1: 文件1的数据帧
            df2: 文件2的数据帧（可选，为None时使用单表比较）
            
        返回:
            tuple: (passed_rules, failed_rules, all_failed_cells, all_passed_cells)
                - passed_rules: 通过的规则列表
                - failed_rules: 失败的规则列表
                - all_failed_cells: 所有失败的单元格列表，格式为[(rule, row_idx, col_idx), ...]
                - all_passed_cells: 所有通过的单元格列表，格式为[(rule, row_idx, col_idx), ...]
        """
        if df2 is None:
            logger.info(f"使用单表模式验证规则，df1形状: {df1.shape}")
        else:
            logger.info(f"使用双表模式验证规则，df1形状: {df1.shape}, df2形状: {df2.shape}")
        return self.rule_engine.validate_with_dataframes(df1, df2)
    
    def compare_with_rules(self, alias1, sheet_name1, alias2=None, sheet_name2=None):
        """
        基于自定义规则比较数据
        
        参数:
            alias1: 第一个工作簿别名
            sheet_name1: 第一个工作表名称
            alias2: 第二个工作簿别名（可选，用于跨文件比较）
            sheet_name2: 第二个工作表名称（可选，用于跨文件比较）
            
        返回:
            tuple: (result_summary, comparison_results)
                - result_summary: 比较结果摘要，包含通过和失败的规则数量
                - comparison_results: 详细比较结果，包含每条规则的验证结果
        """
        logger.info(f"开始基于规则比较数据")
        logger.info(f"文件1: {alias1}，工作表: {sheet_name1}")
        if alias2 and sheet_name2:
            logger.info(f"文件2: {alias2}，工作表: {sheet_name2}")
        
        try:
            # 获取数据
            df1 = self.get_sheet_dataframe(alias1, sheet_name1)
            df2 = self.get_sheet_dataframe(alias2, sheet_name2) if alias2 and sheet_name2 else None
            
            # 验证所有规则
            passed_rules, failed_rules = self.rule_engine.validate_all_rules(df1, df2)
            
            # 计算结果
            total_rules = len(passed_rules) + len(failed_rules)
            passed_rate = len(passed_rules) / total_rules if total_rules > 0 else 1.0
            
            # 生成结果摘要
            result_summary = {
                'total_rules': total_rules,
                'passed_rules': len(passed_rules),
                'failed_rules': len(failed_rules),
                'passed_rate': passed_rate
            }
            
            # 生成详细比较结果
            comparison_results = {
                'passed': passed_rules,
                'failed': failed_rules
            }
            
            logger.info(f"规则比较完成: 总规则数={total_rules}，通过={len(passed_rules)}，失败={len(failed_rules)}，通过率={passed_rate:.2f}")
            if passed_rules:
                logger.info(f"通过的规则: {passed_rules}")
            if failed_rules:
                logger.info(f"失败的规则: {failed_rules}")
            
            return result_summary, comparison_results
        except Exception as e:
            logger.error(f"规则比较失败: {str(e)}")
            # 返回包含错误信息的结果
            result_summary = {
                'total_rules': 0,
                'passed_rules': 0,
                'failed_rules': 0,
                'passed_rate': 0.0,
                'error': str(e)
            }
            comparison_results = {
                'passed': [],
                'failed': [],
                'error': str(e)
            }
            return result_summary, comparison_results
    
    def compare_sheets_with_rules(self, alias1, sheet_name1, alias2, sheet_name2, cell_range=None):
        """
        比较两个工作表中的指定范围，基于自定义规则
        
        参数:
            alias1: 第一个工作簿别名
            sheet_name1: 第一个工作表名称
            alias2: 第二个工作簿别名
            sheet_name2: 第二个工作表名称
            cell_range: 单元格范围（可选）
            
        返回:
            tuple: (result_summary, comparison_results, combined_df)
                - result_summary: 比较结果摘要
                - comparison_results: 详细比较结果
                - combined_df: 组合数据框，包含两个工作表的数据
        """
        logger.info(f"开始比较两个工作表")
        logger.info(f"文件1: {alias1}，工作表: {sheet_name1}")
        logger.info(f"文件2: {alias2}，工作表: {sheet_name2}")
        if cell_range:
            logger.info(f"比较范围: {cell_range}")
        
        try:
            # 获取两个工作表的数据
            df1 = self.get_sheet_dataframe(alias1, sheet_name1)
            df2 = self.get_sheet_dataframe(alias2, sheet_name2)
            
            # 如果指定了范围，选择指定范围的数据
            if cell_range:
                df1 = self.select_cells(alias1, sheet_name1, cell_range)
                df2 = self.select_cells(alias2, sheet_name2, cell_range)
                logger.info(f"范围选择后，文件1数据形状: {df1.shape}，文件2数据形状: {df2.shape}")
                
                # 直接使用范围选择后的数据进行规则验证
                passed_rules, failed_rules = self.rule_engine.validate_all_rules(df1, df2)
                
                # 计算结果
                total_rules = len(passed_rules) + len(failed_rules)
                passed_rate = len(passed_rules) / total_rules if total_rules > 0 else 1.0
                
                # 生成结果摘要
                result_summary = {
                    'total_rules': total_rules,
                    'passed_rules': len(passed_rules),
                    'failed_rules': len(failed_rules),
                    'passed_rate': passed_rate
                }
                
                # 生成详细比较结果
                comparison_results = {
                    'passed': passed_rules,
                    'failed': failed_rules
                }
                
                logger.info(f"规则比较完成: 总规则数={total_rules}，通过={len(passed_rules)}，失败={len(failed_rules)}，通过率={passed_rate:.2f}")
            else:
                # 验证所有规则
                result_summary, comparison_results = self.compare_with_rules(alias1, sheet_name1, alias2, sheet_name2)
            
            # 创建组合数据框用于显示
            max_rows = max(df1.shape[0], df2.shape[0])
            max_cols = max(df1.shape[1], df2.shape[1])
            
            # 确保列名一致
            col_names = []
            for i in range(max_cols):
                if i < df1.shape[1]:
                    col_names.append(f"文件1_{str(df1.columns[i])}")
                elif i < df2.shape[1]:
                    col_names.append(f"文件2_{str(df2.columns[i])}")
                else:
                    col_names.append(f"COL_{i}")
            
            # 构建组合数据框
            combined_data = []
            for r in range(max_rows):
                row_data = []
                # 添加文件1的数据
                for c in range(df1.shape[1]):
                    if r < df1.shape[0]:
                        row_data.append(df1.iloc[r, c])
                    else:
                        row_data.append(None)
                # 添加文件2的数据
                for c in range(df2.shape[1]):
                    if r < df2.shape[0]:
                        row_data.append(df2.iloc[r, c])
                    else:
                        row_data.append(None)
                combined_data.append(row_data)
            
            combined_df = pd.DataFrame(combined_data, columns=col_names)
            logger.info(f"组合数据框创建完成，形状: {combined_df.shape}")
            
            return result_summary, comparison_results, combined_df
        except Exception as e:
            logger.error(f"工作表比较失败: {str(e)}")
            # 返回包含错误信息的结果
            result_summary = {
                'total_rules': 0,
                'passed_rules': 0,
                'failed_rules': 0,
                'passed_rate': 0.0,
                'error': str(e)
            }
            comparison_results = {
                'passed': [],
                'failed': [],
                'error': str(e)
            }
            # 返回空的组合数据框
            combined_df = pd.DataFrame()
            return result_summary, comparison_results, combined_df
